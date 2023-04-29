# -*- coding: utf-8 -*-

import sys
import os
import shutil
import openpyxl

def main():
	command_dict = {
		'move': shutil.move,
		'copy': shutil.copy,
		'mkdir': os.makedirs
	}

	if len(sys.argv) < 3:
		sys.stderr.write('Usage: cat list python $0 input_xls_filename input_xls_sheetame\n')
		sys.exit(1)
	
	input_xls_filename = sys.argv[1]
	input_xls_sheetname = sys.argv[2]
	
	try:
		wb_input = openpyxl.load_workbook(input_xls_filename, data_only=True)
		sheet_input = wb_input[input_xls_sheetname]
	except:
		sys.stderr.write('Failed to open ' + input_xls_filename + ': ' + input_xls_sheetname + '\n')
		sys.exit(1)
	
	# header should be:
	#  A1: 'Command'
	#  B1: Path_From
	#  C1: Path_To
	header = ['Command', 'Path_From', 'Path_To']
	for idx_col in range(1, 4):
		val = sheet_input.cell(row=1, column=idx_col).value
		if val != header[idx_col-1]:
			sys.stderr.write('Wrong header: ' + val + ' should be ' + header[idx_col-1] + '\n')
			sys.exit(1)

	sheet_input.cell(row=1, column=4).value = 'Done'
	
	# process files
	# input:
	#  A: command ('copy' / 'move')
	#  B: path_from
	#  C: path_to
	flg_eol = 0
	idx_row = 2
	while flg_eol == 0:
		if sheet_input.cell(row=idx_row, column=1).value is None or sheet_input.cell(row=idx_row, column=2).value is None or sheet_input.cell(row=idx_row, column=3).value is None:
			flg_eol = 1
		elif len(sheet_input.cell(row=idx_row, column=1).value) == 0 or len(sheet_input.cell(row=idx_row, column=2).value) == 0 or len(sheet_input.cell(row=idx_row, column=3).value) == 0:
			flg_eol = 1
		elif idx_row > 100000:
			flg_eol = 1
		
		if flg_eol == 0:
			this_command = sheet_input.cell(row=idx_row, column=1).value
			path_from = sheet_input.cell(row=idx_row, column=2).value
			path_to = sheet_input.cell(row=idx_row, column=3).value

			this_func = command_dict[this_command]
			this_func(path_from, path_to)

			sheet_input.cell(row=idx_row, column=4).value = 'x'

		idx_row += 1

	wb_input.save(input_xls_filename)
	wb_input.close()

if __name__ == '__main__':
	main()

