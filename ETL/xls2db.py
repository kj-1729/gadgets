import sys
import os
import openpyxl

def process(wb_data, wb_output):
	labels = ['seqno', 'sheet_name', 'row', 'column', 'datatype', 'value']
	for loop in range(len(labels)):
		wb_output['output'].cell(row=1, column=loop+1).value = labels[loop]

	idx_y_out = 2
	seqno = 0
	for this_sheet in wb_data.sheetnames:
		for idx_y in range(wb_data[this_sheet].min_row, wb_data[this_sheet].max_row+1):
			for idx_x in range(wb_data[this_sheet].min_column, wb_data[this_sheet].max_column+1):
				if wb_data[this_sheet].cell(row=idx_y, column=idx_x).value is not None and len(str(wb_data[this_sheet].cell(row=idx_y, column=idx_x).value)) > 0:
					wb_output['output'].cell(row=idx_y_out, column=1).value = seqno
					wb_output['output'].cell(row=idx_y_out, column=2).value = this_sheet
					wb_output['output'].cell(row=idx_y_out, column=3).value = idx_y
					wb_output['output'].cell(row=idx_y_out, column=4).value = idx_x
					this_value = wb_data[this_sheet].cell(row=idx_y, column=idx_x).value
					wb_output['output'].cell(row=idx_y_out, column=5).value = str(type(this_value))
					
					if type(this_value) is str and this_value[0] == '=':
						this_value = "'" + this_value
					wb_output['output'].cell(row=idx_y_out, column=6).value = this_value

					seqno += 1
					idx_y_out += 1

def main():
	if len(sys.argv) < 3:
		sys.stderr.write('Usage: xls2db.py (datafile) (otuputfile)\n')
		sys.exit(1)
	
	input_path = sys.argv[1]
	output_path = sys.argv[2]
	
	# get input/output excel file
	wb_data = openpyxl.load_workbook(input_path)
	#wb_data = openpyxl.load_workbook(input_path, data_only=True)
	
	wb_output = openpyxl.Workbook()
	wb_output.create_sheet(index=0, title='output')
	
	process(wb_data, wb_output)
	
	wb_output.save(output_path)
	
	wb_data.close()
	wb_output.close()

if __name__ == '__main__':
	main()

