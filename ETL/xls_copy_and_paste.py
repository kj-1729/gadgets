import sys
import os
import openpyxl

def process(sheet_config, wb_from, wb_to):
	flg = 1
	
	idx_row = 9
	while flg > 0:
		sheet_from = sheet_config.cell(row=idx_row, column=1).value
		row_from = sheet_config.cell(row=idx_row, column=2).value
		col_from = sheet_config.cell(row=idx_row, column=3).value
		sheet_to = sheet_config.cell(row=idx_row, column=4).value
		row_to = sheet_config.cell(row=idx_row, column=5).value
		col_to = sheet_config.cell(row=idx_row, column=6).value
		
		wb_to[sheet_to].cell(row=row_to, column=col_to).value = wb_from[sheet_from].cell(row=row_from, column=col_from).value

		idx_row += 1
		if sheet_config.cell(row=idx_row, column=1).value is None:
			flg = 0
		elif len(sheet_config.cell(row=idx_row, column=1).value) == 0:
			flg = 0
		elif idx_row > 100000:
			flg = 0

def main():
	if len(sys.argv) < 2:
		sys.stderr.write('Usage: copy_and_paste.py config_file\n')
		sys.exit(1)
	
	config_path = sys.argv[1]
	
	# get input/output excel file
	wb_config = openpyxl.load_workbook(config_path, data_only=True)
	sheet_config = wb_config['config']
	dir_from = sheet_config.cell(row=2, column=4).value
	fname_from = sheet_config.cell(row=3, column=4).value
	dir_to = sheet_config.cell(row=4, column=4).value
	fname_to = sheet_config.cell(row=5, column=4).value
	path_from = os.path.join(dir_from, fname_from)
	path_to = os.path.join(dir_to, fname_to)
	
	wb_from = openpyxl.load_workbook(path_from, data_only=True)
	wb_to = openpyxl.load_workbook(path_to)
	
	process(sheet_config, wb_from, wb_to)
	
	wb_to.save(path_to)
	
	wb_config.close()
	wb_from.close()
	wb_to.close()

if __name__ == '__main__':
	main()

